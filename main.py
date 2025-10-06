import os
import re
import docx
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox

class DocxToExcelAutomator:
    def __init__(self, root):
        self.root = root
        self.root.title("Inova Process - DOCX to Excel Automator")
        self.root.geometry("500x200")

        self.file_path = tk.StringVar()

        # --- UI Elements ---

        # Frame for file selection
        file_frame = tk.Frame(root)
        file_frame.pack(pady=10, padx=10, fill=tk.X)

        # Label for file selection
        self.label = tk.Label(file_frame, text="Select the Bill of Materials file (.docx):")
        self.label.pack(anchor=tk.W)

        # Entry box for file path
        self.path_entry = tk.Entry(file_frame, textvariable=self.file_path, state='readonly', width=60)
        self.path_entry.pack(side=tk.LEFT, expand=True, fill=tk.X, ipady=2)

        # Browse button
        self.browse_button = tk.Button(file_frame, text="Browse...", command=self.browse_file)
        self.browse_button.pack(side=tk.RIGHT, padx=(5, 0))

        # Start Automation button
        self.start_button = tk.Button(root, text="Start Automation", command=self.start_automation, font=("Helvetica", 12, "bold"))
        self.start_button.pack(pady=20, padx=10, fill=tk.X, ipady=5)

        # Status label
        self.status_label = tk.Label(root, text="Status: Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)

    def browse_file(self):
        """Opens a file dialog to select a .docx file."""
        filepath = filedialog.askopenfilename(
            title="Select a DOCX file",
            filetypes=(("Word Documents", "*.docx"), ("All files", "*.*"))
        )
        if filepath:
            self.file_path.set(filepath)
            self.update_status(f"Selected: {os.path.basename(filepath)}")

    def start_automation(self):
        """Starts the main automation process."""
        arquivo_word = self.file_path.get()
        if not arquivo_word:
            messagebox.showerror("Error", "Please select a .docx file first.")
            return

        # Assuming the Excel file is in the same directory and has a fixed name
        planilha_excel = os.path.join(os.path.dirname(arquivo_word), "TABELA-DE-AÇO R8.xlsx")

        if not os.path.exists(planilha_excel):
            messagebox.showerror("Error", f"Excel file not found:\n{planilha_excel}")
            return

        self.update_status("Processing...")
        self.root.update_idletasks() # Force UI update

        try:
            dados_extraidos = self.extrair_dados_word(arquivo_word)
            if dados_extraidos:
                self.preencher_planilha_excel(planilha_excel, dados_extraidos)
                self.update_status("Success: Excel sheet updated.")
                messagebox.showinfo("Success", "The Excel sheet has been successfully updated.")
            else:
                self.update_status("No data extracted from Word file.")
                messagebox.showwarning("Warning", "No data was extracted from the Word file. Please check the file.")
        except Exception as e:
            self.update_status(f"Error: {e}")
            messagebox.showerror("Automation Error", f"An error occurred: {e}")


    def update_status(self, message):
        """Updates the status label text."""
        self.status_label.config(text=f"Status: {message}")

    # MÓDULO DE INTELIGÊNCIA DE ENGENHARIA
    #______________________________________________________________________
    def convert_to_mm(self, dim_str):
        """Converte dimensões em polegadas (ex: "1.1/2"") para mm."""
        dim_str = dim_str.strip().replace(',', '.')
        total_mm = 0.0
        try:
            if '"' in dim_str:
                dim_str = dim_str.replace('"', '')
                parts = dim_str.split('.')
                if parts[0] and '/' in parts[0]:
                    num, den = map(float, parts[0].split('/'))
                    total_mm += (num / den) * 25.4
                elif parts[0]:
                    total_mm += float(parts[0]) * 25.4
                if len(parts) > 1 and '/' in parts[1]:
                    num, den = map(float, parts[1].split('/'))
                    total_mm += (num / den) * 25.4
            else:
                total_mm = float(dim_str)
        except (ValueError, ZeroDivisionError): return 0.0
        return total_mm

    # ==============================================================================
    # Função para identificar o tipo de perfil e retornar o código correto
    # ==============================================================================
    def classificar_e_mapear_perfil(self, desc):
        """Identifica o TIPO de perfil e retorna o código e uma chave de classificação."""
        desc_upper = desc.upper()
        if '[' in desc_upper or '][' in desc_upper: return 'U.s', 'PERFIL_U'
        if 'UENR' in desc_upper or 'IENR' in desc_upper or 'CART' in desc_upper or 'CA ' in desc_upper: return 'U.e', 'TERCA'
        if 'L DOBRADO' in desc_upper or desc_upper.startswith('L '): return 'L DOBRADO', 'CANTONEIRA'


        # Se encontrar 'RED', mapeia para o código exato da planilha.
        if 'RED' in desc_upper:
            return 'FERRO MECANICO RED.', 'TUBO' # Mapeamento corrigido

        if 'TUBO' in desc_upper: return 'TUBO', 'TUBO' # Mantém genérico se for outra coisa

        return 'N/D', 'OUTROS'


    # ==============================================================================
    # Extrai as 4 medidas principais de uma descrição de perfil
    # ==============================================================================
    def parse_dimensoes_inteligente(self, desc, tipo_perfil):
        """Aplica regras de extração de dimensões e retorna as 4 medidas principais."""
        a, b, c, esp = 0.0, 0.0, 0.0, 0.0
        numeros_str_list = re.findall(r'[\d\./"]+', desc)

        if tipo_perfil in ['PERFIL_U']:
            if len(numeros_str_list) >= 3:
                a = self.convert_to_mm(numeros_str_list[0])
                b = self.convert_to_mm(numeros_str_list[1])
                esp = self.convert_to_mm(numeros_str_list[2])
        elif tipo_perfil == 'TERCA':
            if len(numeros_str_list) >= 4:
                a = self.convert_to_mm(numeros_str_list[0])
                b = self.convert_to_mm(numeros_str_list[1])
                c = self.convert_to_mm(numeros_str_list[2])
                esp = self.convert_to_mm(numeros_str_list[3])
        elif tipo_perfil == 'CANTONEIRA':
            if len(numeros_str_list) == 2:
                aba = self.convert_to_mm(numeros_str_list[0])
                a, b = aba, aba
                esp = self.convert_to_mm(numeros_str_list[1])
            elif len(numeros_str_list) >= 3:
                a = self.convert_to_mm(numeros_str_list[0])
                b = self.convert_to_mm(numeros_str_list[1])
                esp = self.convert_to_mm(numeros_str_list[2])

        # --- CORREÇÃO APLICADA AQUI ---
        # Lógica específica para Tubo/RED
        elif tipo_perfil == 'TUBO':
            # Para RED 12.7, a única medida é a espessura/diâmetro.
            if len(numeros_str_list) >= 1:
                esp = self.convert_to_mm(numeros_str_list[0]) # Coloca o valor na variável 'esp'

        return a, b, c, esp

    # ==============================================================================
    # MÓDULO PRINCIPAL DO SCRIPT (Leitura e Preenchimento Não-Destrutivo)
    # ==============================================================================

    def extrair_dados_word(self, caminho_arquivo_word):
        """Função robusta para extrair dados do Word."""
        documento = docx.Document(caminho_arquivo_word)
        tabela = documento.tables[0]
        if len(tabela.rows) < 2: return None
        perfils_str = tabela.cell(1, 0).text; acos_str = tabela.cell(1, 1).text
        ltotais_str = tabela.cell(1, 2).text; pesos_str = tabela.cell(1, 3).text
        lista_perfis = list(filter(None, perfils_str.strip().split('\n')))
        lista_acos = list(filter(None, acos_str.strip().split('\n')))
        lista_ltotais = list(filter(None, ltotais_str.strip().split('\n')))
        lista_pesos = list(filter(None, pesos_str.strip().split('\n')))
        num_perfis = len(lista_perfis)
        if not (num_perfis == len(lista_ltotais) == len(lista_pesos)): return None
        if num_perfis == 0: return None
        
        dados_finais = []
        for i in range(num_perfis):
            perfil, aco = lista_perfis[i].strip(), lista_acos[i].strip() if i < len(lista_acos) else lista_acos[0].strip()
            l_total_str, peso_str = lista_ltotais[i].strip().replace(',', '.'), lista_pesos[i].strip().replace(',', '.')
            try:
                l_total_m = float(l_total_str) / 100 if l_total_str else 0.0
                peso_final = float(peso_str) if peso_str else 0.0
                dados_finais.append([perfil, aco, l_total_m, peso_final])
            except ValueError: continue
        return dados_finais

    def encontrar_proxima_linha_vazia(self, sheet, codigo_secao, linha_inicio_busca):
        """
        Encontra a primeira linha vazia para uma seção, aceitando placeholders como 'X' ou 0.
        """
        for row in range(linha_inicio_busca, sheet.max_row + 2):
            celula_codigo = sheet.cell(row=row, column=1)
            celula_dado_ref = sheet.cell(row=row, column=2)
            if celula_codigo.value == codigo_secao and celula_dado_ref.value in [None, 0, 'X', '']:
                return row
        return None

    def preencher_planilha_excel(self, caminho_planilha, dados_materiais):
        """Preenche a planilha de forma não-destrutiva, seguindo a estrutura de colunas exata."""
        workbook = openpyxl.load_workbook(caminho_planilha)
        sheet = workbook.active
        
        dados_agrupados = {}
        for item in dados_materiais:
            codigo_excel, _ = self.classificar_e_mapear_perfil(item[0])
            if codigo_excel not in dados_agrupados: dados_agrupados[codigo_excel] = []
            dados_agrupados[codigo_excel].append(item)
            
        for codigo_secao, itens_da_secao in dados_agrupados.items():
            linha_de_busca_secao = 4
            for item in itens_da_secao:
                linha_alvo = self.encontrar_proxima_linha_vazia(sheet, codigo_secao, linha_de_busca_secao)
                
                if linha_alvo is None:
                    print(f"  AVISO: Não há mais espaço na planilha para a seção '{codigo_secao}'. Item '{item[0]}' não inserido.")
                    continue

                perfil_desc, aco_tipo, l_total_m, peso_total = item
                _, tipo_perfil = self.classificar_e_mapear_perfil(perfil_desc)
                dim_a, dim_b, dim_c, dim_esp = self.parse_dimensoes_inteligente(perfil_desc, tipo_perfil)

                if tipo_perfil in ['PERFIL_U', 'TERCA']:
                    sheet.cell(row=linha_alvo, column=2).value = dim_a
                    sheet.cell(row=linha_alvo, column=4).value = dim_b
                    sheet.cell(row=linha_alvo, column=6).value = dim_c
                elif tipo_perfil == 'CANTONEIRA':
                    sheet.cell(row=linha_alvo, column=4).value = dim_a
                    sheet.cell(row=linha_alvo, column=6).value = dim_b

                sheet.cell(row=linha_alvo, column=8).value = dim_esp
                sheet.cell(row=linha_alvo, column=9).value = aco_tipo
                sheet.cell(row=linha_alvo, column=10).value = l_total_m
                sheet.cell(row=linha_alvo, column=17).value = peso_total
                linha_de_busca_secao = linha_alvo + 1

        workbook.save(caminho_planilha)

# ==============================================================================
# PONTO DE PARTIDA DO SCRIPT
# ==============================================================================
if __name__ == "__main__":
    root = tk.Tk()
    app = DocxToExcelAutomator(root)
    root.mainloop()
